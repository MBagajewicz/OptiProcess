#!/usr/bin/env python3
"""Initialize/update OptiAppsCreator users from an Excel file.

Expected columns: username, email, password
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from auth_db import connect, get_db_path, init_db, upsert_user, utc_iso


REQUIRED_COLUMNS = {"username", "email", "password"}


def main() -> None:
    parser = argparse.ArgumentParser(description="Import users from Excel into the OptiHexx auth database")
    parser.add_argument("--file", required=True, help="Excel file path with username, email, password columns")
    parser.add_argument("--db", default=None, help="Optional SQLite DB path. Defaults to AUTH_DB_PATH/data/users.db")
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="Create only new users; keep credentials for matching usernames or email addresses",
    )
    args = parser.parse_args()

    excel_path = Path(args.file)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    db_path = Path(args.db) if args.db else get_db_path()
    init_db(db_path)

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, ())
    normalized = {str(value).strip().lower(): index for index, value in enumerate(headers) if value is not None}
    missing = REQUIRED_COLUMNS - set(normalized)
    if missing:
        workbook.close()
        raise SystemExit(f"Missing required columns: {', '.join(sorted(missing))}")

    count = 0
    skipped = 0
    direct_access_enabled = 0
    for row in rows:
        username_value = row[normalized["username"]]
        email_value = row[normalized["email"]]
        password_value = row[normalized["password"]]
        username = str(username_value).strip() if username_value is not None else ""
        email = str(email_value).strip().lower() if email_value is not None else ""
        password = str(password_value) if password_value is not None else ""
        if not username or not email or not password:
            continue
        if args.skip_existing:
            with connect(db_path) as conn:
                existing = conn.execute(
                    """
                    SELECT id, initial_password_hash, must_change_password
                    FROM users
                    WHERE username = ? OR email = ?
                    """,
                    (username, email),
                ).fetchone()
                if existing and (existing["must_change_password"] or existing["initial_password_hash"]):
                    conn.execute(
                        """
                        UPDATE users
                        SET initial_password_hash = NULL, must_change_password = 0, updated_at = ?
                        WHERE id = ?
                        """,
                        (utc_iso(), existing["id"]),
                    )
                    direct_access_enabled += 1
            if existing:
                skipped += 1
                continue
        upsert_user(username, email, password, db_path=db_path)
        count += 1
    workbook.close()

    print(
        f"Imported {count} users into {db_path}; skipped {skipped} existing users; "
        f"enabled direct access for {direct_access_enabled} existing users"
    )


if __name__ == "__main__":
    main()
