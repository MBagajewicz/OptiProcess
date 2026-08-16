#region Title: SimulatorResultHFM
#Nature: Results generator
#Methodology: Uses results of simulation and generate output methods to be used by HFM-Simulator
##################################################################################################################
#VERSION        DATE            AUTHOR                    DESCRIPTION OF CHANGES MADE
#0.0       13-May-2025    Diego Gabriel Oliva            Commented
#0.0       24-Oct-2025    Dego Oliva                     Unified Mass & Energy in a single Excel sheet, added 0-filling for missing energy data
#0.2       24-Oct-2025    Dego Oliva                     Removed top metadata rows. Added Case, Discretizations, and Components as repeated columns at the start.
##################################################################################################################
#endregion

import os
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from Common.Stream.stream import Stream
from Common.Membrane_Properties.Permeance.Membrane_Permeance import MembranePermeance

class SimulatorResultsHFM:
    """
    Container for all simulation results.
    Stores geometry, mass model variables, and energy model variables.
    Provides export utilities and component-based accessors.
    """

    def __init__(self):
        # -----------------------------
        # geometry & general
        # -----------------------------
        self.NCells: int = 0
        self.z: np.ndarray = np.array([])
        self.components: list = []
        self.case_name: tuple = ("simulation", "case")

        # -----------------------------
        # mass model
        # -----------------------------
        self.FRet: np.ndarray = np.array([])
        self.FPerm: np.ndarray = np.array([])
        self.ZRet: np.ndarray = np.array([])
        self.ZPerm: np.ndarray = np.array([])
        self.PRetCell: np.ndarray = np.array([])
        self.PPermCell: np.ndarray = np.array([])
        self.FMemb: np.ndarray = np.array([])
        self.FMemb_comp: np.ndarray = np.array([])
        self.ZMemb: np.ndarray = np.array([])
        self.FugacityRet: np.ndarray = np.array([])
        self.FugacityPerm: np.ndarray = np.array([])

        # -----------------------------
        # energy model
        # -----------------------------
        self.T_ret: np.ndarray = np.array([])
        self.T_per: np.ndarray = np.array([])
        self.hRet: np.ndarray = np.array([])
        self.hPerm: np.ndarray = np.array([])
        self.hMemb: np.ndarray = np.array([])
        self.UA: np.ndarray = np.array([])
        self.U: np.ndarray = np.array([])

        # Ladder rungs that produced solutions ('LU', 'fixed_point', 'marching',
        # 'least_squares'). The fast paths fail silently and fall through, so
        # this is the only record that a result came from the last-resort solver.
        self.solver_paths: list = []

        # Dew-point condition, decided by the phase-stability test in the
        # simulator. This is what the enumeration constraint must read; the
        # Tdew_* arrays below come from the PQ root-find and are for reporting.
        self.dew_ok: bool = True
        self.dew_bad_node = None
        self.dew_bad_side = None
        self.Tdew_ret: np.ndarray = np.array([])
        self.Tdew_per: np.ndarray = np.array([])
        self.Tdew_mem: np.ndarray = np.array([])

        # Physical Properties & Membrane
        self.Permeance = None
        self.viscosity = None
        self.molecularweight = None

        # Save T feed when energy balance is False
        self.T_feed = None
    # ==========================================
    # DERIVED QUANTITIES & ACCESSORS
    # ==========================================

    @property
    def recovery(self):
        """Fraction of feed recovered in permeate."""
        if self.FRet is None or len(self.FRet) == 0:
            return None
        return 1 - self.FRet[-1] / self.FRet[0]

    def _comp_index(self, comp):
        """Resolve component index from name or int."""
        if isinstance(comp, int):
            if comp < 0 or comp >= len(self.components):
                raise IndexError(f"Component index {comp} out of range (0-{len(self.components)-1})")
            return comp
        if isinstance(comp, str):
            if comp not in self.components:
                raise ValueError(f"Component '{comp}' not found. Available: {self.components}")
            return self.components.index(comp)
        raise TypeError("Component must be name (str) or index (int)")

    def component_flux(self, comp):
        """Membrane flux profile for a component."""
        return self.FMemb_comp[:, self._comp_index(comp)]

    def retentate_composition(self, comp):
        """Retentate composition profile for a component."""
        return self.ZRet[:, self._comp_index(comp)]

    def permeate_composition(self, comp):
        """Permeate composition profile for a component."""
        return self.ZPerm[:, self._comp_index(comp)]

    def component_retentate_flow(self, comp):
        """Component flow rate profile in retentate."""
        return self.FRet * self.ZRet[:, self._comp_index(comp)]

    def component_permeate_flow(self, comp):
        """Component flow rate profile in permeate."""
        return self.FPerm * self.ZPerm[:, self._comp_index(comp)]

    def list_components(self):
        """Return list of components."""
        return self.components

    def outlet(self, side="retentate"):
        """Return outlet stream as a new Stream object."""
        if side == "retentate":
            comp_dict = {
                comp: float(frac)
                for comp, frac in zip(self.components, self.ZRet[-1])
            }
            T = (
                float(self.T_ret[-1])
                if len(self.T_ret) > 0
                else (self.T_feed if self.T_feed is not None else 298.15)
            )
            return Stream(
                composition=comp_dict,
                P=float(self.PRetCell[-1]),
                T=T,
                molar_flow=float(self.FRet[-1]),
            )

        if side == "permeate":
            comp_dict = {
                comp: float(frac)
                for comp, frac in zip(self.components, self.ZPerm[0])
            }
            T = (
                float(self.T_per[0])
                if len(self.T_per) > 0
                else (self.T_feed if self.T_feed is not None else 298.15)
            )
            return Stream(
                composition=comp_dict,
                P=float(self.PPermCell[0]),
                T=T,
                molar_flow=float(self.FPerm[0]),
            )

        raise ValueError("side must be 'retentate' or 'permeate'")

    # ==========================================
    # EXCEL EXPORT HELPERS (Boundary handling)
    # ==========================================

    def _get_ret_flow(self, k, i):
        return round(float(self.FRet[k] * self.ZRet[k, i]), 6)

    def _get_perm_flow(self, k, i):
        return round(float(self.FPerm[k] * self.ZPerm[k, i]), 6) if k < self.NCells else ""

    def _get_ret_fug(self, k, i):
        return float(self.FugacityRet[k, i]) if k > 0 else ""

    def _get_perm_fug(self, k, i):
        return float(self.FugacityPerm[k, i]) if k < self.NCells else ""

    def _get_perm_tot(self, k):
        return round(float(self.FPerm[k]), 6) if k < self.NCells else ""

    def _get_perm_comp(self, k, i):
        return float(self.ZPerm[k, i]) if k < self.NCells else ""

    # ==========================================
    # HEADER & ROW BUILDERS
    # ==========================================

    def _build_header(self, n_comp: int) -> list:
        """Builds the unified header row."""
        header = ["Case", "Discretizations", "Components", "Elapsed Time [s]", "z [m]", "F_tot [mol/s]"]

        # Mass: Retentate
        for i in range(n_comp): header.append(f"FRet_{i}")
        for i in range(n_comp): header.append(f"ZRet[{i}]")
        header.append("PRet [Pa]")
        for i in range(n_comp): header.append(f"FugacityRet[{i}] [Pa]")

        # Mass: Permeate
        header.append("FPerm_tot [mol/s]")
        for i in range(n_comp): header.append(f"FPerm_{i}")
        for i in range(n_comp): header.append(f"ZPerm[{i}]")
        header.append("PPerm [Pa]")
        for i in range(n_comp): header.append(f"FugacityPerm[{i}] [Pa]")

        # Mass: Membrane
        header.append("FMemb_tot [mol/s]")
        for i in range(n_comp): header.append(f"FMemb_{i}")
        for i in range(n_comp): header.append(f"ZMemb[{i}]")

        # Energy
        header += ["hRet", "hPerm", "hMemb", "T_ret", "T_per", "Q_cond",
                    "Tdew_ret", "Tdew_per", "Tdew_mem"]

        return header

    def _build_row(self, k: int, n_comp: int, N: int, case_str: str, disc_str, comp_str: str, has_energy: bool, elapsed_time) -> list:
        """Builds a single data row for node k."""
        z_val = round(float(self.z[k]), 6) if self.z is not None and len(self.z) > k else k

        row = [case_str, disc_str, comp_str, elapsed_time, z_val, round(float(self.FRet[k]), 6)]

        # 1. Retentate Mass Data
        for i in range(n_comp): row.append(self._get_ret_flow(k, i))
        for i in range(n_comp): row.append(float(self.ZRet[k, i]))
        row.append(float(self.PRetCell[k]))
        for i in range(n_comp): row.append(self._get_ret_fug(k, i))

        # 2. Permeate Mass Data
        row.append(self._get_perm_tot(k))
        for i in range(n_comp): row.append(self._get_perm_flow(k, i))
        for i in range(n_comp): row.append(self._get_perm_comp(k, i))
        row.append(float(self.PPermCell[k]))
        for i in range(n_comp): row.append(self._get_perm_fug(k, i))

        # 3. Membrane Mass Data
        row.append(float(self.FMemb[k]))
        for i in range(n_comp): row.append(float(self.FMemb_comp[k, i]))
        for i in range(n_comp): row.append(float(self.ZMemb[k, i]))

        # 4. Energy Data
        if has_energy:
            h_ret = float(self.hRet[k])
            h_perm = float(self.hPerm[k])
            h_memb = float(self.hMemb[k])
            t_ret = float(self.T_ret[k])
            t_per = float(self.T_per[k]) if k < N else ""
            q_cond = float(self.UA[k] * (self.T_ret[k] - self.T_per[k-1])) if k > 0 else ""
            tdew_ret = float(self.Tdew_ret[k])
            tdew_per = float(self.Tdew_per[k])
            tdew_mem = float(self.Tdew_mem[k])
        else:
            h_ret = h_perm = h_memb = t_ret = t_per = q_cond = 0.0
            tdew_ret = tdew_per = tdew_mem = 0.0

        row += [h_ret, h_perm, h_memb, t_ret, t_per, q_cond, tdew_ret, tdew_per, tdew_mem]

        return row

    # ==========================================
    # UNIFIED EXCEL EXPORT (Single result)
    # ==========================================

    def export_results_to_excel(
        self,
        filename: str = "HFM_Results.xlsx",
        case_name: str = "Default_Case",
        append: bool = False,  # <--- CAMBIO CLAVE: Ahora por defecto es False (sobrescribe/pisa el archivo)
        elapsed_time: any = ""
    ) -> None:
        """
        Exports simulation results to a single Excel worksheet.

        Parameters
        ----------
        append : bool
            If False (default), overwrites the file if it exists.
            If True, appends rows to the existing file (used for batch/validator runs).
        """
        n_comp = len(self.components)
        N = self.NCells
        has_energy = self.T_ret is not None and len(self.T_ret) > 0

        case_str = str(case_name)
        disc_str = N
        comp_str = ", ".join(self.components)

        time_str = f"{elapsed_time:.2f}" if isinstance(elapsed_time, (int, float)) else elapsed_time

        file_exists = os.path.isfile(filename)

        # Lógica de sobrescritura vs append
        if append and file_exists:
            wb = load_workbook(filename)
            ws = wb.active
            write_header = False
        else:
            # Si append=False, simplemente creamos un Workbook nuevo.
            # Al hacer wb.save() más abajo, openpyxl pisará el archivo existente automáticamente.
            wb = Workbook()
            ws = wb.active
            ws.title = "HFM_Results"
            write_header = True

        if write_header:
            ws.append(self._build_header(n_comp))

        for k in range(N + 1):
            row = self._build_row(k, n_comp, N, case_str, disc_str, comp_str, has_energy, time_str)
            ws.append(row)

        wb.save(filename)
        action = "Appended to" if (append and file_exists) else "Overwritten/Created"
        print(f"✅ {action}: {filename}  (Case: {case_name}, Nodes: {N+1})")

    # ==========================================
    # BATCH EXPORT (Multiple results at once)
    # ==========================================

    @staticmethod
    def export_batch(
        results_list: list,
        case_names: list,
        filename: str = "HFM_Results.xlsx",
        overwrite: bool = True
    ) -> None:
        """
        Exports multiple SimulatorResultsHFM objects to a single Excel file.
        """
        if len(results_list) != len(case_names):
            raise ValueError(f"results_list ({len(results_list)}) and case_names ({len(case_names)}) must have the same length.")

        if overwrite and os.path.isfile(filename):
            os.remove(filename)
            print(f"🗑️  Overwritten: {filename}")

        for results, name in zip(results_list, case_names):
            results.export_results_to_excel(
                filename=filename,
                case_name=name,
                append=True  # <--- El batch SIEMPRE usa append=True internamente
            )

        total_rows = sum(r.NCells + 1 for r in results_list)
        print(f"\n🎉 Batch export complete: {len(results_list)} cases, {total_rows} total rows → {filename}")